import io
import json

import openpyxl
from fastapi import Form, File, UploadFile
from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import HTMLResponse
from starlette.responses import JSONResponse
from datetime import datetime, timedelta
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse
from crud import CRUDUsers
from crud.dialogCRUD import CRUDDialog
import pandas as pd
import os
from urllib.parse import unquote

from schemas.userSchemas import UserSchemaExel

templates = Jinja2Templates(directory="templates")
router = APIRouter(
    prefix='/admin_view',
    tags=['Admin View']
)


@router.get("/today")
async def statistic_today(request: Request):
    # today = datetime.now().strftime("%Y-%m-%d")
    gradeUser = 0
    gradeAdmin = 0

    allDialogs = await CRUDDialog.get_all_td()

    allDialogsLen = len(await CRUDDialog.get_all_td())
    openDialogs = len(await CRUDDialog.get_all_today(is_active=True))
    closeDialogs = len(await CRUDDialog.get_all_today(is_active=False))

    for getGrade in allDialogs:
        gradeUser += getGrade.gradeUser
        gradeAdmin += getGrade.gradeAdmin

    try:
        averageGradeUser = round(gradeUser / allDialogsLen, 1)
        averageGradeAdmin = round(gradeAdmin / allDialogsLen, 1)
    except ZeroDivisionError:
        averageGradeUser = 0
        averageGradeAdmin = 0

    context = {
        "allDialogs": allDialogsLen,
        "openDialogs": openDialogs,
        "closeDialogs": closeDialogs,
        "gradeUser": averageGradeUser,
        "gradeAdmin": averageGradeAdmin,
        "nameDays": "today"
    }

    return JSONResponse(content=context)


@router.get("/week")
async def statistic_week(request: Request):
    # today = datetime.now()
    # start_of_week = today - timedelta(days=today.weekday())
    # end_of_week = start_of_week + timedelta(days=6)

    gradeUser = 0
    gradeAdmin = 0

    allDialogs = await CRUDDialog.get_all_wk()

    allDialogsLen = len(await CRUDDialog.get_all_wk())
    openDialogs = len(await CRUDDialog.get_all_week(is_active=True))
    closeDialogs = len(await CRUDDialog.get_all_week(is_active=False))
    for getGrade in allDialogs:
        gradeUser += getGrade.gradeUser
        gradeAdmin += getGrade.gradeAdmin
    try:
        averageGradeUser = round(gradeUser / allDialogsLen, 1)
        averageGradeAdmin = round(gradeAdmin / allDialogsLen, 1)
    except ZeroDivisionError:
        averageGradeUser = 0
        averageGradeAdmin = 0

    context = {
        "allDialogs": allDialogsLen,
        "openDialogs": openDialogs,
        "closeDialogs": closeDialogs,
        "gradeUser": averageGradeUser,
        "gradeAdmin": averageGradeAdmin,
        "nameDays": "week"
    }

    return JSONResponse(content=context)


@router.get("/month")
async def statistic_month(request: Request):
    # today = datetime.now()
    # start_of_month = today.replace(day=1)
    # end_of_month = (today.replace(day=1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    gradeUser = 0
    gradeAdmin = 0

    allDialogs = await CRUDDialog.get_all_mn()

    allDialogsLen = len(await CRUDDialog.get_all_mn())
    openDialogs = len(await CRUDDialog.get_all_month(is_active=True))
    closeDialogs = len(await CRUDDialog.get_all_month(is_active=False))
    for getGrade in allDialogs:
        gradeUser += getGrade.gradeUser
        gradeAdmin += getGrade.gradeAdmin
    try:
        averageGradeUser = round(gradeUser / allDialogsLen, 1)
        averageGradeAdmin = round(gradeAdmin / allDialogsLen, 1)
    except ZeroDivisionError:
        averageGradeUser = 0
        averageGradeAdmin = 0

    context = {
        "allDialogs": allDialogsLen,
        "openDialogs": openDialogs,
        "closeDialogs": closeDialogs,
        "gradeUser": averageGradeUser,
        "gradeAdmin": averageGradeAdmin,
        "nameDays": "month"
    }

    return JSONResponse(content=context)


@router.get("/test")
async def statistic_week(request: Request):
    getIsBlockUser = len(await CRUDUsers.get_all_is_block(is_block=True))
    getIsBlockUserFalse = len(await CRUDUsers.get_all_is_block(is_block=False))

    inside_count, outside_count = await CRUDUsers.get_count_today()
    context = {
        "getIsBlockUser": getIsBlockUser,
        "getIsBlockUserFalse": getIsBlockUserFalse,
        "inside_count": inside_count,
        "outside_count": outside_count,
    }
    return JSONResponse(content=context)


@router.get("/download/{name}/{data}")
async def download(name: str, data: str):
    data = json.loads(unquote(data))
    getAllDialogs = [data["getAllDialogs"]]
    openDialogs = [data["openDialogs"]]
    closeDialogs = [data["closeDialogs"]]
    gradeUser = [data["gradeUser"]]
    gradeAdmin = [data["gradeAdmin"]]
    getName = [name]
    df = pd.DataFrame({
        'Когда': getName,
        'Всего диалогов': getAllDialogs,
        'Открыто диалогов': openDialogs,
        'Закрыто': closeDialogs,
        'Средняя оценка Продавца': gradeUser,
        'Средняя оценка Саппорта': gradeAdmin,
    })
    df.to_excel('Statistics.xlsx')

    file_path = "Statistics.xlsx"
    if os.path.exists(file_path):
        return FileResponse(file_path,
                            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            filename="Statistics.xlsx")
    else:
        raise HTTPException(status_code=404, detail="File not found")


@router.post("/add_user", response_class=JSONResponse)
async def add_user(request: Request, file: UploadFile = File(...)):
    try:
        # Чтение данных из временного файла и создание объекта io.BytesIO
        file_content = await file.read()
        file_io = io.BytesIO(file_content)

        # Используем load_workbook без контекстного менеджера
        file_to_read = openpyxl.load_workbook(file_io, data_only=True)
        sheet = file_to_read['id']

        users_added = 0
        no_users_added = 0

        for row in range(2, sheet.max_row + 1):
            data = []
            for col in range(1, 6):
                value = sheet.cell(row, col).value
                data.append(value)
            user_exists = await CRUDUsers.get_lnr_phone(lnr=str(data[3]), phone=str(data[4]))
            if not user_exists:
                await CRUDUsers.add_in_exel(user=UserSchemaExel(
                    last_name=data[0],
                    first_name=data[1],
                    middle_name=data[2],
                    lnr=str(data[3]),
                    phone=str(data[4])
                ))
                users_added += 1
            else:
                no_users_added += 1

        if no_users_added:
            message = f"Добавлено {users_added} пользователей\n\n" \
                      f"Не добавлено {no_users_added} пользователей"
        else:
            message = f"Добавлено {users_added} пользователей"

        file_to_read.close()

        return {"status": "success", "message": message}
    except Exception as e:
        return {"status": "error", "message": f"Произошла ошибка: {e}"}

